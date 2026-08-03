from PyQt6.QtCore import QEvent, QObject, Qt, QTimer, QDate
from PyQt6.QtGui import QGuiApplication, QKeySequence
from PyQt6.QtWidgets import (QAbstractItemView, QHBoxLayout, QHeaderView,
                             QLabel, QMenu, QPushButton, QTableWidget, QTableWidgetItem,
                             QVBoxLayout, QWidget, QFileDialog, QMessageBox)

from libs.CalendarLineEdit import DateRangeLineEdit
from libs.CompleterLineEdit import CompleterLineEdit
from libs.CustomSpinBox import CustomSpinBox
from libs.DatabaseConnector import DatabaseConnector
from libs.EditHistory import EditHistoryDialog

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ReadOnlyTable(QTableWidget):
    def __init__(self, headers: list[str], parent=None):
        super().__init__(parent)
        self.headers = headers

        # Read-only and selection behavior
        self.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.setSelectionMode(QAbstractItemView.SelectionMode.ContiguousSelection)
        self.setAlternatingRowColors(True)
        self.setSortingEnabled(True)
        self.verticalHeader().setVisible(False)

        # Disable manual resizing
        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)

        # Context menu
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)

        # Columns and headers
        self.setColumnCount(len(headers))
        self.setHorizontalHeaderLabels(headers)
        self.horizontalHeader().setStretchLastSection(True)

        # Enable Ctrl+C for copying selected rows
        self.installEventFilter(self)

    def show_context_menu(self, pos):
        item = self.itemAt(pos)
        if item:
            menu = QMenu(self)
            print_action = menu.addAction("Edit     ")
            if menu.exec(self.mapToGlobal(pos)) == print_action:
                row = item.row()
                values = [self.item(row, col).text() for col in range(self.columnCount())]
                data = dict(zip(self.headers, values))
                self.edit_history = EditHistoryDialog(data=data, parent=self)
                self.edit_history.exec()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.KeyPress and event.matches(QKeySequence.StandardKey.Copy):
            self.copy_selection_to_clipboard()
            return True
        return super().eventFilter(obj, event)

    def copy_selection_to_clipboard(self):
        selected_ranges = self.selectedRanges()
        if not selected_ranges:
            return
        copied_text = ""
        # Add headers (from left to right column of the first selected range)
        first_range = selected_ranges[0]
        headers = [
            self.horizontalHeaderItem(col).text()
            for col in range(first_range.leftColumn(), first_range.rightColumn() + 1)
        ]
        copied_text += "\t".join(headers) + "\n"
        # Add selected data
        for selection in selected_ranges:
            for row in range(selection.topRow(), selection.bottomRow() + 1):
                row_data = [
                    self.item(row, col).text() if self.item(row, col) else ""
                    for col in range(selection.leftColumn(), selection.rightColumn() + 1)
                ]
                copied_text += "\t".join(row_data) + "\n"
        QGuiApplication.clipboard().setText(copied_text.strip())

    def mouseDoubleClickEvent(self, event):
        item = self.itemAt(event.pos())
        if item:
            row = item.row()
            self.clearSelection()
            self.selectRow(row)
        super().mousePressEvent(event)


class History(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.database = DatabaseConnector()

        self.headers = [
            "ID", "BHW Name", "Date Replaced", "Run Count", "SAP#",
            "Qty. of Pogo Pins Replaced", "Total Price in Euro",
            "Site/s", "Replaced by", "Remarks"
        ]

        self.table = ReadOnlyTable(self.headers)

        # Limit with label
        self.limit = CustomSpinBox(width=200, value=100, parent=self)
        self.limit.valueChanged.connect(self.schedule_reload)
        limit_label = QLabel("Limit:")
        limit_label.setMaximumWidth(100)
        limit_label.setProperty("role", "historyLabel")
        self.limit.setProperty("role", "historyInput")

        # BHW data input with label
        completer_lb_lst = self.database.get_all_lb()
        self.bhw_data = CompleterLineEdit(completer_lb_lst, 200, self.load_bhw_history, enter_func=True, parent=self)
        bhw_label = QLabel("BHW Name:")
        bhw_label.setMaximumWidth(100)
        bhw_label.setProperty("role", "historyLabel")
        self.bhw_data.setProperty("role", "historyInput")

        # Date range with label
        self.date_range = DateRangeLineEdit(func=self.load_by_date, parent=self)
        date_label = QLabel("Date Range:")
        date_label.setMaximumWidth(100)
        date_label.setProperty("role", "historyLabel")
        self.date_range.setProperty("role", "historyInput")

        # Export button
        self.export_button = QPushButton("Export to Excel")
        self.export_button.setFixedWidth(150)
        self.export_button.setProperty("role", "historyExport")
        self.export_button.clicked.connect(self.export_to_excel)
        
        # Clear Filter button
        self.clear_filter_button = QPushButton("Clear Filter")
        self.clear_filter_button.setProperty("role", "historyClrFilter")
        self.clear_filter_button.setFixedWidth(100)
        self.clear_filter_button.clicked.connect(self.clear_filters)

        # Controls container layout: add all labels and inputs directly
        controls_layout = QHBoxLayout()
        controls_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)

        controls_layout.addWidget(limit_label)
        controls_layout.addWidget(self.limit)

        controls_layout.addWidget(bhw_label)
        controls_layout.addWidget(self.bhw_data)

        controls_layout.addWidget(date_label)
        controls_layout.addWidget(self.date_range)

        # Add buttons to controls layout
        controls_layout.addWidget(self.export_button)
        controls_layout.addWidget(self.clear_filter_button)

        main_layout = QVBoxLayout(self)
        main_layout.addWidget(self.table)
        main_layout.addLayout(controls_layout)

        self.installEventFilter(self)

        self.reload_timer = QTimer(self)
        self.reload_timer.setSingleShot(True)
        self.reload_timer.timeout.connect(self.reload_table_data)

        self.load_all_history()

    def eventFilter(self, obj: QObject, event: QEvent) -> bool:
        if obj is self and event.type() == QEvent.Type.MouseButtonPress:
            pos = event.position().toPoint()
            widget = self.childAt(pos)
            if not widget or (widget is not self.table and not self.table.isAncestorOf(widget)):
                self.table.clearSelection()
        return super().eventFilter(obj, event)

    def schedule_reload(self):
        self.reload_timer.start(2000)  # Restart 2-second timer

    def reload_table_data(self):
        self.clear_table()
        self.load_all_history()

    def load_all_history(self):
        data = self.database.get_all_history(self.limit.value())
        self.load_data(data, update_limit=False)

    def load_data(self, data: list[tuple], update_limit=True):
        """
        Load data into the table.
        
        Args:
            data: List of tuples containing the data
            update_limit: If True, update the limit spinbox to show the count
        """
        # Ensure data is not empty and has consistent length
        if not data:
            self.table.setRowCount(0)
            if update_limit:
                self.limit.blockSignals(True)
                self.limit.setValue(0)
                self.limit.blockSignals(False)
            return

        # Determine column count from the first row
        column_count = len(data[0])
        self.table.setColumnCount(column_count)
        self.table.setRowCount(len(data))

        # Set table headers
        if hasattr(self, 'headers'):
            self.table.setHorizontalHeaderLabels(self.headers)

        for row_idx, row_data in enumerate(data):
            for col_idx, col_data in enumerate(row_data):
                # Special handling for ID column (column 0) - store as integer for proper sorting
                if col_idx == 0:
                    try:
                        numeric_value = int(col_data)
                        item = QTableWidgetItem()
                        item.setData(Qt.ItemDataRole.DisplayRole, numeric_value)
                        item.setData(Qt.ItemDataRole.EditRole, numeric_value)
                    except (ValueError, TypeError):
                        item = QTableWidgetItem("0")
                        item.setData(Qt.ItemDataRole.DisplayRole, 0)
                        item.setData(Qt.ItemDataRole.EditRole, 0)
                else:
                    # For other columns, use string display
                    display_value = "" if col_data in (None, "", " ") else str(col_data)
                    item = QTableWidgetItem(display_value)
                
                item.setFlags(Qt.ItemFlag.ItemIsSelectable | Qt.ItemFlag.ItemIsEnabled)
                self.table.setItem(row_idx, col_idx, item)

        # Update limit spinbox to show actual record count if requested
        if update_limit:
            self.limit.blockSignals(True)
            self.limit.setValue(len(data))
            self.limit.blockSignals(False)

        # Resize all columns except the last
        for col in range(column_count - 1):
            self.table.resizeColumnToContents(col)

        # Stretch the last column after layout update
        QTimer.singleShot(0, lambda: self.table.horizontalHeader().setStretchLastSection(True))

    def clear_table(self):
        self.table.setRowCount(0)

    def load_bhw_history(self, bhw: str, command=None):
        self.clear_table()
        bhw_history = self.database.get_bhw_history(bhw, command)
        if bhw_history:
            QTimer.singleShot(100, lambda: self.load_data(bhw_history, update_limit=False))
        else:
            self.limit.blockSignals(True)
            self.limit.setValue(0)
            self.limit.blockSignals(False)

    def load_by_date(self):
        """
        Load records filtered by date range.
        The limit is ignored and the spinbox shows the actual count.
        """
        # Defensive check for empty or malformed text
        text = self.date_range.text()
        if not text or " - " not in text:
            return
        
        try:
            start_date, end_date = text.split(" - ")
            # Strip any whitespace
            start_date = start_date.strip()
            end_date = end_date.strip()
            
            data_in_range = self.database.get_bhw_history_in_range(start_date, end_date)
            self.clear_table()
            
            # Load data and update limit to show actual count
            if data_in_range:
                QTimer.singleShot(100, lambda: self.load_data(data_in_range, update_limit=True))
            else:
                # No data found, set limit to 0
                self.limit.blockSignals(True)
                self.limit.setValue(0)
                self.limit.blockSignals(False)
                
        except Exception as e:
            print(f"Error loading by date: {e}")
            QMessageBox.warning(
                self,
                "Error",
                f"Error loading data by date:\n{str(e)}"
            )

    def clear_filters(self):
        """
        Clear all filters (BHW name and date range) and reload all history.
        """
        # Clear BHW input
        self.bhw_data.clear()
        
        # Clear date range
        self.date_range.clear()
        self.date_range.setPlaceholderText("Select date range")
        
        # Reset limit to default (100)
        self.limit.blockSignals(True)
        self.limit.setValue(100)
        self.limit.blockSignals(False)
        
        # Reload all history
        self.clear_table()
        self.load_all_history()
        
        # Show feedback (optional)
        QMessageBox.information(
            self,
            "Filters Cleared",
            "All filters have been cleared.\nShowing all history records."
        )

    def get_date_range_from_text(self):
        """
        Extract start and end dates from the date range text.
        Returns tuple (start_date, end_date) or (None, None) if invalid.
        """
        text = self.date_range.text()
        if not text or " - " not in text:
            return None, None
        
        try:
            start_str, end_str = text.split(" - ")
            start_date = QDate.fromString(start_str.strip(), "yyyy-MM-dd")
            end_date = QDate.fromString(end_str.strip(), "yyyy-MM-dd")
            
            if start_date.isValid() and end_date.isValid():
                return start_date, end_date
        except:
            pass
        
        return None, None

    def get_current_filters(self):
        """
        Get the current filter state.
        Returns a dict with filter information.
        """
        filters = {
            'bhw': self.bhw_data.text().strip(),
            'date_range': self.date_range.text().strip(),
            'limit': self.limit.value()
        }
        return filters

    def export_to_excel(self):
        """Export the current table data to an Excel file with all data as text/string format"""
        if not OPENPYXL_AVAILABLE:
            QMessageBox.warning(
                self,
                "Export Error",
                "openpyxl is not installed. Please install it using:\npip install openpyxl"
            )
            return

        # Get current data from the table
        row_count = self.table.rowCount()
        col_count = self.table.columnCount()
        
        if row_count == 0:
            QMessageBox.information(
                self,
                "Export Error",
                "No data to export. Please load some data first."
            )
            return

        # Get current filters
        filters = self.get_current_filters()
        bhw_filter = filters['bhw']
        date_range_text = filters['date_range']
        
        # Get date range for filename
        start_date, end_date = self.get_date_range_from_text()
        
        # Generate filename based on filters
        today = QDate.currentDate()
        filename_parts = ["History"]
        
        if bhw_filter:
            filename_parts.append(f"BHW_{bhw_filter}")
        
        if start_date and end_date and start_date.isValid() and end_date.isValid():
            filename_parts.append(f"{start_date.toString('yyyy-MM-dd')}_to_{end_date.toString('yyyy-MM-dd')}")
        else:
            filename_parts.append(f"Export_{today.toString('yyyy-MM-dd')}")
        
        filename = "_".join(filename_parts) + ".xlsx"

        # Ask user where to save the file
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Excel File",
            filename,
            "Excel Files (*.xlsx);;All Files (*)"
        )

        if not file_path:
            return  # User cancelled

        try:
            # Create workbook and get active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "History"

            # Set up styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
            
            cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            # Write headers (as strings)
            for col_idx, header in enumerate(self.headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = str(header)  # Force as string
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # Set cell number format to text
                cell.number_format = '@'

            # Write data - all as strings
            for row_idx in range(row_count):
                for col_idx in range(col_count):
                    item = self.table.item(row_idx, col_idx)
                    value = item.text() if item else ""
                    
                    cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                    cell.value = str(value)  # Force as string
                    cell.alignment = cell_alignment
                    cell.border = border
                    # Set cell number format to text to prevent automatic conversion
                    cell.number_format = '@'

            # Auto-adjust column widths
            for col_idx in range(1, col_count + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                # Check header length
                header_value = ws.cell(row=1, column=col_idx).value
                if header_value:
                    max_length = max(max_length, len(str(header_value)))
                
                # Check data length
                for row_idx in range(2, row_count + 2):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                # Set column width with min and max limits
                adjusted_width = min(max(max_length + 2, 12), 50)  # Min 12, Max 50
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add metadata/info about the export (as strings)
            info_row = row_count + 3
            
            # Export date
            cell = ws.cell(row=info_row, column=1)
            cell.value = f"Export Date: {today.toString('yyyy-MM-dd HH:mm:ss')}"
            cell.font = Font(italic=True)
            cell.number_format = '@'  # Force as text
            
            # Add filter information
            row_offset = 1
            
            if bhw_filter:
                cell = ws.cell(row=info_row + row_offset, column=1)
                cell.value = f"BHW Filter: {bhw_filter}"
                cell.font = Font(italic=True)
                cell.number_format = '@'
                row_offset += 1
            
            if start_date and end_date and start_date.isValid() and end_date.isValid():
                cell = ws.cell(row=info_row + row_offset, column=1)
                cell.value = f"Date Range: {start_date.toString('yyyy-MM-dd')} to {end_date.toString('yyyy-MM-dd')}"
                cell.font = Font(italic=True)
                cell.number_format = '@'
                row_offset += 1
            
            # Add total row count
            cell = ws.cell(row=info_row + row_offset, column=1)
            cell.value = f"Total Records: {row_count}"
            cell.font = Font(italic=True)
            cell.number_format = '@'

            # Save the file
            wb.save(file_path)
            
            QMessageBox.information(
                self,
                "Export Successful",
                f"Data successfully exported to:\n{file_path}\n\n"
                f"Total Records: {row_count}\n"
                f"Filters Applied: {filters}"
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"An error occurred while exporting:\n{str(e)}"
            )

    def get_current_data(self):
        """Helper method to get current table data as list of lists"""
        data = []
        for row_idx in range(self.table.rowCount()):
            row_data = []
            for col_idx in range(self.table.columnCount()):
                item = self.table.item(row_idx, col_idx)
                row_data.append(item.text() if item else "")
            data.append(row_data)
        return data